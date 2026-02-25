"""
Excel multi-sheet export for assessment reports.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Dict, Any, List
from datetime import datetime

def generate_excel_report(assessment_data: Dict[str, Any], comparison_data: List[Dict[str, Any]] = None) -> BytesIO:
    """
    Generate a multi-sheet Excel workbook with assessment details.
    
    Args:
        assessment_data: Single assessment data dict
        comparison_data: Optional list of comparison assessments
    
    Returns:
        BytesIO buffer containing the Excel file
    """
    
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Sheet 1: Assessment Summary
    ws_summary = wb.create_sheet("Assessment Summary", 0)
    _populate_summary_sheet(ws_summary, assessment_data)
    
    # Sheet 2: Design Specification (NEW)
    ws_design = wb.create_sheet("Design Specification", 1)
    _populate_design_sheet(ws_design, assessment_data)
    
    # Sheet 3: Physical & Financial
    ws_physical = wb.create_sheet("Physical & Cost", 2)
    _populate_physical_sheet(ws_physical, assessment_data)
    
    # Sheet 4: Amenities
    ws_amenities = wb.create_sheet("Amenities", 3)
    _populate_amenities_sheet(ws_amenities, assessment_data)
    
    # Sheet 5: Comparison (if provided)
    if comparison_data:
        ws_compare = wb.create_sheet("Comparison", 4)
        _populate_comparison_sheet(ws_compare, comparison_data)
    
    # Sheet 6: Recommendations & Constraints
    ws_recs = wb.create_sheet("Recommendations", 5)
    _populate_recommendations_sheet(ws_recs, assessment_data)

    # Sheet 6: Regulatory Findings
    ws_reg = wb.create_sheet("Regulatory Findings", 5)
    _populate_regulatory_sheet(ws_reg, assessment_data)
    
    # Write to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


def _populate_summary_sheet(ws, assessment_data: Dict[str, Any]):
    """Fill the Assessment Summary sheet."""
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35
    
    # Header
    ws["A1"] = "UR HAPPY HOME - SITE ASSESSMENT SUMMARY"
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1F7F4C", end_color="1F7F4C", fill_type="solid")
    ws.merge_cells("A1:B1")
    
    # Generated date
    ws["A2"] = "Generated:"
    ws["B2"] = datetime.now().strftime("%d %B %Y %H:%M")
    
    # Site info
    row = 4
    headers = ["Field", "Value"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F7F4C", end_color="1F7F4C", fill_type="solid")
    
    row = 5
    data = [
        ("Address", assessment_data.get("address", "")),
        ("Coordinates", f"{assessment_data.get('latitude', 0):.4f}, {assessment_data.get('longitude', 0):.4f}"),
        ("Planning Zone", assessment_data.get("zone_type", "N/A")),
        ("Viability Score", f"{assessment_data.get('raw_score', 0):.1f}/100"),
        ("Status", assessment_data.get("viability_status", "Unknown")),
        ("Transport Distance", f"{assessment_data.get('dist_transport', 0)}m"),
        ("Lot Width", f"{assessment_data.get('lot_width', 0)}m"),
        ("Lot Depth", f"{assessment_data.get('lot_depth', 0)}m"),
        ("Lot Area", f"{assessment_data.get('lot_area', 0):.0f} sqm"),
        ("Heritage Overlay", "YES" if assessment_data.get("has_overlay") else "NO"),
        ("Single Dwelling Covenant", "YES" if assessment_data.get("has_covenant") else "NO"),
    ]
    
    for field, value in data:
        ws.cell(row=row, column=1, value=field).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1


def _populate_design_sheet(ws, assessment_data: Dict[str, Any]):
    """Fill the Design Specification sheet with locked design details."""
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 50
    
    # Header
    ws["A1"] = "UR HAPPY HOME STANDARD DESIGN SPECIFICATION"
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1F7F4C", end_color="1F7F4C", fill_type="solid")
    ws.merge_cells("A1:B1")
    
    row = 3
    ws.cell(row=row, column=1, value="Design Details").font = Font(bold=True, size=11)
    row += 1
    
    design_data = [
        ("Design Name", assessment_data.get("design_locked", "UR Happy Home Standard Design v1.0")),
        ("Status", "APPROVED - February 2026"),
        ("Compliance", "Rooming-house standards + NDIS regulations"),
        ("", ""),
        ("Bedrooms", "5"),
        ("Bathrooms", "3 (inc. 1 ensuite)"),
        ("Gross Floor Area", "274 m²"),
        ("Building Levels", "2"),
        ("Site Footprint", "9.74m × 9.74m (~95 m²)"),
        ("", ""),
        ("Kitchens", "2"),
        ("Dining/Meals Areas", "2"),
        ("Living Areas", "2"),
        ("", ""),
        ("Accessible Bedrooms", "3 fully accessible"),
        ("Accessible Bathrooms", "2"),
        ("Wheelchair Access", "1200mm corridors minimum"),
        ("Doorways", "850mm minimum width"),
        ("Handrails/Grab Rails", "Yes"),
        ("Emergency Exits", "2 compliant exits"),
        ("", ""),
        ("Heating System", "Heat pump (all-electric ready)"),
        ("Cooling System", "Ducted reverse-cycle"),
        ("Power Supply", "3-phase for commercial kitchen"),
        ("Hot Water", "Heat pump system"),
        ("EV Charging Ready", "Yes"),
        ("Solar Ready", "Yes"),
        ("", ""),
        ("Planning Permit", "Required (GFA > 300 m²)"),
        ("Exemptions Applicable", "No (5 bedrooms exceeds 3-bedroom exemption limit)"),
        ("Max Occupancy", "5 persons (< 12 limit) ✓"),
    ]
    
    for field, value in design_data:
        if field == "":
            row += 1
        else:
            ws.cell(row=row, column=1, value=field).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
    
    row += 2
    ws.cell(row=row, column=1, value="Site Suitability").font = Font(bold=True, size=11)
    row += 1
    
    suit = assessment_data.get("design_suitability") or {}
    if suit and "error" not in suit:
        suit_data = [
            ("Overall Suitability", "Suitable ✓" if suit.get("all_checks_pass") else "Requires Review"),
            ("Lot Area Check", "Pass" if suit.get("suitability_checks", {}).get("lot_area_sufficient") else "Review"),
            ("Zone Suitable", "Yes" if suit.get("suitability_checks", {}).get("zone_suitable") else "Confirm"),
            ("Transport Access", "Good" if suit.get("suitability_checks", {}).get("transport_compliant") else "Note"),
            ("Title Check Required", "Yes"),
        ]
        for field, value in suit_data:
            ws.cell(row=row, column=1, value=field).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
    else:
        ws.cell(row=row, column=1, value="Suitability assessment not available")
        row += 1


def _populate_physical_sheet(ws, assessment_data: Dict[str, Any]):
    """Fill the Physical & Cost sheet."""
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35
    
    ws["A1"] = "PHYSICAL SUITABILITY & COST ESTIMATE"
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1F7F4C", end_color="1F7F4C", fill_type="solid")
    ws.merge_cells("A1:B1")
    
    row = 3
    headers = ["Metric", "Value"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F7F4C", end_color="1F7F4C", fill_type="solid")
    
    row = 4
    data = [
        ("Lot Width", f"{assessment_data.get('lot_width', 0)}m (≥14m)"),
        ("Lot Depth", f"{assessment_data.get('lot_depth', 0)}m (≥24m)"),
        ("Lot Area", f"{assessment_data.get('lot_area', 0):.0f} sqm (≥336 sqm)"),
        ("Site Slope", assessment_data.get("slope", "Unknown")),
        ("Regulatory: Fixed Heating", "✓" if assessment_data.get("check_heating") else "✗"),
        ("Regulatory: Blind Cord Safety", "✓" if assessment_data.get("check_windows") else "✗"),
        ("Regulatory: All-Electric Ready", "✓" if assessment_data.get("check_energy") else "✗"),
    ]
    
    for field, value in data:
        ws.cell(row=row, column=1, value=field).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1


def _populate_amenities_sheet(ws, assessment_data: Dict[str, Any]):
    """Fill the Amenities sheet."""
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    
    ws["A1"] = "NEARBY AMENITIES (1km RADIUS)"
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1F7F4C", end_color="1F7F4C", fill_type="solid")
    ws.merge_cells("A1:B1")
    
    amenities = assessment_data.get("amenities_summary", {})
    row = 3
    
    categories = [
        ("Transit", amenities.get("transit", [])),
        ("Schools", amenities.get("schools", [])),
        ("Parks", amenities.get("parks", [])),
        ("Shops", amenities.get("shops", [])),
        ("Heritage", amenities.get("heritage", [])),
    ]
    
    for category, items in categories:
        ws.cell(row=row, column=1, value=category).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        row += 1
        
        for item in items[:5]:  # Top 5 per category
            ws.cell(row=row, column=1, value=item.get("name", ""))
            ws.cell(row=row, column=2, value=f"{item.get('distance_m', 0)}m")
            row += 1
        
        row += 1


def _populate_comparison_sheet(ws, comparison_data: List[Dict[str, Any]]):
    """Fill the Comparison sheet."""
    ws.column_dimensions["A"].width = 20
    for col in range(2, len(comparison_data) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    ws["A1"] = "SITE COMPARISON"
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1F7F4C", end_color="1F7F4C", fill_type="solid")
    
    # Headers
    headers = ["Metric"] + [f"Site {i+1}" for i in range(len(comparison_data))]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F7F4C", end_color="1F7F4C", fill_type="solid")
    
    # Comparison metrics
    metrics = [
        ("Address", lambda d: d.get("address")),
        ("Score", lambda d: f"{d.get('raw_score', 0):.1f}"),
        ("Status", lambda d: d.get("viability_status")),
        ("Zone", lambda d: d.get("zone_type")),
        ("Lot Area (sqm)", lambda d: f"{d.get('lot_area', 0):.0f}"),
        ("Transport (m)", lambda d: f"{d.get('dist_transport', 0)}"),
    ]
    
    row = 3
    for metric, extractor in metrics:
        ws.cell(row=row, column=1, value=metric).font = Font(bold=True)
        for col, data in enumerate(comparison_data, 2):
            ws.cell(row=row, column=col, value=extractor(data))
        row += 1


def _populate_recommendations_sheet(ws, assessment_data: Dict[str, Any]):
    """Fill the Recommendations & Constraints sheet."""
    ws.column_dimensions["A"].width = 40
    
    ws["A1"] = "RECOMMENDATIONS & CONSTRAINTS"
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1F7F4C", end_color="1F7F4C", fill_type="solid")
    ws.merge_cells("A1:A1")
    
    # Constraints
    ws["A2"] = "IDENTIFIED CONSTRAINTS"
    ws["A2"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A2"].fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    
    row = 3
    constraints = assessment_data.get("identified_constraints", [])
    if constraints:
        for constraint in constraints:
            ws.cell(row=row, column=1, value=f"• {constraint}").alignment = Alignment(wrap_text=True)
            row += 1
    else:
        ws.cell(row=row, column=1, value="No significant constraints identified")
        row += 1
    
    row += 1
    
    # Recommendations
    ws.cell(row=row, column=1, value="RECOMMENDATIONS").font = Font(bold=True, size=11, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill(start_color="1F7F4C", end_color="1F7F4C", fill_type="solid")
    row += 1
    
    recommendations = assessment_data.get("recommendations", [])
    for rec in recommendations:
        ws.cell(row=row, column=1, value=f"• {rec}").alignment = Alignment(wrap_text=True)
        row += 1
    
    row += 1
    ws.cell(row=row, column=1, value="NEXT STEPS").font = Font(bold=True, size=11, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill(start_color="1F7F4C", end_color="1F7F4C", fill_type="solid")
    row += 1
    
    next_steps = [
        "Conduct detailed legal due diligence (title search, covenants, encumbrances)",
        "Obtain formal builder/construction quotes",
        "Engage town planning consultant for design review",
        "Confirm with council regarding development approval pathway",
        "Commission geotechnical survey if slope > 5%",
        "Arrange property valuation for financing"
    ]
    
    for step in next_steps:
        ws.cell(row=row, column=1, value=f"• {step}").alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1


def _populate_regulatory_sheet(ws, assessment_data: Dict[str, Any]):
    """Fill the Regulatory Findings sheet with rooming-house standard checks."""
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 80

    ws['A1'] = 'REGULATORY FINDINGS - ROOMING-HOUSE MINIMUM STANDARDS'
    ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='1F7F4C', end_color='1F7F4C', fill_type='solid')
    ws.merge_cells('A1:B1')

    row = 3
    reg = assessment_data.get('regulatory_findings') or {}
    standards = reg.get('standards') or {}

    ws.cell(row=row, column=1, value='Standard').font = Font(bold=True)
    ws.cell(row=row, column=2, value='Value / Notes').font = Font(bold=True)
    row += 1

    if standards:
        ws.cell(row=row, column=1, value='Max gross floor area (sqm)')
        ws.cell(row=row, column=2, value=str(standards.get('max_gross_floor_area_sqm')))
        row += 1
        ws.cell(row=row, column=1, value='Max persons accommodated')
        ws.cell(row=row, column=2, value=str(standards.get('max_persons_accommodated')))
        row += 1
        ws.cell(row=row, column=1, value='Max bedrooms')
        ws.cell(row=row, column=2, value=str(standards.get('max_bedrooms')))
        row += 2

    ws.cell(row=row, column=1, value='Check').font = Font(bold=True)
    ws.cell(row=row, column=2, value='Result / Details').font = Font(bold=True)
    row += 1

    checks = reg.get('checks', {})
    for k, v in checks.items():
        ws.cell(row=row, column=1, value=k)
        ws.cell(row=row, column=2, value=('PASS' if v else 'FAIL'))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='Detailed reasons').font = Font(bold=True)
    row += 1
    for reason in reg.get('reasons', []):
        ws.cell(row=row, column=1, value=reason)
        row += 1
